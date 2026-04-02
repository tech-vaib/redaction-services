#pip install google-cloud-dlp pandas openpyxl pdfplumber fpdf2
#export GOOGLE_APPLICATION_CREDENTIALS="/path/to/your-service-account.json"
#########
How to run:
1. For Raw Text
python google_dlp_deidentify.py -t "John Doe's email is john@example.com"
2.For a JSON File
python google_dlp_deidentify.py -i data.json
3. For CSV
python google_dlp_deidentify.py -i sales_data.csv
4.For Excel
python google_dlp_deidentify.py -i report.xlsx
5.For PDF
python google_dlp_deidentify.py -i document.pdf

##############################
import os
import csv
import json
import pdfplumber
import openpyxl
import tempfile

from fpdf import FPDF
from google.cloud import dlp_v2

# ====================
# Configuration
# ====================

# Replace with your GCP project ID
GOOGLE_PROJECT_ID = os.getenv("GOOGLE_PROJECT_ID") or "<YOUR_PROJECT_ID>"

# InfoTypes to scan for
INFO_TYPES = [
    {"name": "PERSON_NAME"},
    {"name": "EMAIL_ADDRESS"},
    {"name": "PHONE_NUMBER"},
    {"name": "CREDIT_CARD_NUMBER"},
    {"name": "US_SOCIAL_SECURITY_NUMBER"},
    {"name": "DATE"},
]

# ====================
# Google DLP Client
# ====================

dlp_client = dlp_v2.DlpServiceClient()
parent = f"projects/{GOOGLE_PROJECT_ID}"

# ====================
# DLP Deidentify API
# ====================

def deidentify_text(text):
    """Use Google DLP to deidentify text."""
    #inspect_config = {"info_types": INFO_TYPES, "min_likelihood": dlp_v2.Likelihood.POSSIBLE}
     inspect_config = {"info_types": [], "min_likelihood": dlp_v2.Likelihood.POSSIBLE}
     #build in +custom
     inspect_config = {"info_types": [],  # all built-in
            "custom_info_types": [{"name": "INTERNAL_ORDER_ID","regex": {"pattern": r"\bORD\d{6}\b"}}]}
    deidentify_config = {
        "info_type_transformations": {
            "transformations": [
                {
                    "info_types": INFO_TYPES,
                    "primitive_transformation": {
                        "replace_with_info_type_config": {}
                    }
                }
            ]
        }
    }

    response = dlp_client.deidentify_content(
        request={
            "parent": parent,
            "inspect_config": inspect_config,
            "deidentify_config": deidentify_config,
            "item": {"value": text},
        }
    )
    return response.item.value

# ====================
# JSON Handler
# ====================

def process_json_file(input_path, output_path):
    with open(input_path, "r") as f:
        data = json.load(f)

    def recursive_deid(obj):
        if isinstance(obj, str):
            return deidentify_text(obj)
        if isinstance(obj, dict):
            return {k: recursive_deid(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [ recursive_deid(x) for x in obj ]
        return obj

    deid = recursive_deid(data)

    with open(output_path, "w") as f:
        json.dump(deid, f, indent=2)
    print(f"[INFO] JSON redacted saved to {output_path}")

# ====================
# CSV Handler
# ====================

def process_csv_file(input_path, output_path):
    rows = []
    with open(input_path, "r") as f:
        reader = csv.reader(f)
        header = next(reader)
        for row in reader:
            rows.append(row)

    # Deidentify per field
    new_rows = []
    for row in rows:
        new_row = [ deidentify_text(cell) for cell in row ]
        new_rows.append(new_row)

    with open(output_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(new_rows)
    print(f"[INFO] CSV redacted saved to {output_path}")

# ====================
# Excel Handler
# ====================

def process_xlsx_file(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell.value = deidentify_text(cell.value)

    wb.save(output_path)
    print(f"[INFO] XLSX redacted saved to {output_path}")

# ====================
# PDF Handler
# ====================

def process_pdf_file(input_path, output_path):
    """Extract text from PDF, deidentify, and re-build a PDF."""
    pdf = pdfplumber.open(input_path)
    all_text = ""

    for page in pdf.pages:
        all_text += page.extract_text() + "\n"

    pdf.close()

    redacted_text = deidentify_text(all_text)

    pdf_out = FPDF()
    pdf_out.add_page()
    pdf_out.set_font("Arial", size=12)
    for line in redacted_text.split("\n"):
        pdf_out.cell(200, 10, txt=line, ln=1)
    pdf_out.output(output_path)
    print(f"[INFO] PDF redacted saved to {output_path}")

# ====================
# MAIN PROCESS
# ====================

def process_file(input_path):
    ext = input_path.split(".")[-1].lower()

    # Build output file name
    base = os.path.basename(input_path)
    name = os.path.splitext(base)[0]
    out = f"{name}_redacted.{ext}"

    if ext == "json":
        process_json_file(input_path, out)
    elif ext == "csv":
        process_csv_file(input_path, out)
    elif ext in ("xlsx", "xls"):
        process_xlsx_file(input_path, out)
    elif ext == "pdf":
        process_pdf_file(input_path, out)
    else:
        print(f"[WARN] Unsupported format: {ext}")

# ====================
# RUN FROM COMMAND LINE
# ====================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Google DLP Redaction Tool")
    parser.add_argument("--input", "-i", required=True, help="Path to input file or text")
    parser.add_argument("--is_text", "-t", default=False, action="store_true", help="Treat input as raw text")
    args = parser.parse_args()

    if args.is_text:
        print("[INFO] Deidentifying raw text input …")
        print(deidentify_text(args.input))
    else:
        print(f"[INFO] Processing file: {args.input}")
        process_file(args.input)
